import requests
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# Substitua pela sua API KEY da OpenWeather
API_KEY = "64181ae1e7ac2214229b906341dfba00"
CIDADE = "Sao Paulo"
URL = f"http://api.openweathermap.org/data/2.5/weather?q={CIDADE}&appid={API_KEY}&lang=pt_br&units=metric"

# Nome do arquivo Excel
ARQUIVO = "clima.xlsx"

# Função para salvar os dados
def salvar_dados(temp, umidade):
    agora = datetime.now()
    data_hora = agora.strftime("%d/%m/%Y %H:%M:%S")

    # Se o arquivo não existir, crie
    if not os.path.exists(ARQUIVO):
        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico Clima"
        ws.append(["Data/Hora", "Temperatura (°C)", "Umidade (%)"])
    else:
        wb = load_workbook(ARQUIVO)
        ws = wb.active

    ws.append([data_hora, temp, umidade])
    wb.save(ARQUIVO)

# Função principal para buscar e salvar o clima
def buscar_clima():
    try:
        resposta = requests.get(URL)
        dados = resposta.json()

        if resposta.status_code == 200:
            temperatura = dados['main']['temp']
            umidade = dados['main']['humidity']
            salvar_dados(temperatura, umidade)

            messagebox.showinfo("Sucesso", f"✔️ Dados salvos!\nTemperatura: {temperatura}°C\nUmidade: {umidade}%")
        else:
            messagebox.showerror("Erro", f"Erro ao buscar dados: {dados.get('message', 'Desconhecido')}")

    except Exception as erro:
        messagebox.showerror("Erro", f"Ocorreu um erro: {erro}")

# Interface Tkinter
janela = tk.Tk()
janela.title("Captador de Temperatura - São Paulo")
janela.geometry("400x200")
janela.configure(bg="#e0f7fa")

titulo = tk.Label(janela, text="Captador de Temperatura de São Paulo", font=("Arial", 14), bg="#e0f7fa")
titulo.pack(pady=20)

botao = tk.Button(janela, text="Buscar previsão", command=buscar_clima, font=("Arial", 12), bg="#00838f", fg="white", padx=10, pady=5)
botao.pack()

janela.mainloop()
